#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
file_reader_service.py — 统一文件读取服务

参照 link_shanghai-1.py 的核心逻辑，专门读取：
  1. shanghaidata_x.xlsx  — 节点坐标数据（332行，BD09坐标）
  2. link_matrix_shanghai.csv — 链接矩阵（332×332，浮点权重）

输入（命令行参数）:
  --mode=all      读取全部文件（默认）
  --mode=xlsx     仅读取 xlsx 坐标
  --mode=csv      仅读取 csv 链接矩阵
  --xlsx=path     xlsx 文件路径（可选，有默认值）
  --csv=path      csv 文件路径（可选，有默认值）

输出（stdout JSON）:
  {
    "xlsx": { "success":bool, "nodes":[...], "nodeCount":int, "columns":[...], ... },
    "csv":  { "success":bool, "edges":[...], "edgeCount":int, "matrixShape":[n,m], ... },
    "errors": [str, ...]
  }
"""

import sys
import json
import csv
import os
import argparse
import time
import math

# ============================================================
# 默认文件路径（相对于项目根目录）
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
DEFAULT_XLSX = os.path.join(PROJECT_ROOT, "ai_models_flow", "shanghai_data", "shanghaidata_x.xlsx")
DEFAULT_CSV = os.path.join(PROJECT_ROOT, "ai_models_flow", "shanghai_data", "generate_acc_shanghai", "link_matrix_shanghai.csv")

# ============================================================
# 坐标转换函数（与 link_shanghai-1.py 一致）
# ============================================================

def bd09_to_gcj02(bd_lon, bd_lat):
    """BD09 → GCJ02 坐标转换"""
    x = bd_lon - 0.0065
    y = bd_lat - 0.006
    z = math.sqrt(x * x + y * y) - 0.00002 * math.sin(y * math.pi * 3000.0 / 180.0)
    theta = math.atan2(y, x) - 0.000003 * math.cos(x * math.pi * 3000.0 / 180.0)
    gcj_lon = z * math.cos(theta)
    gcj_lat = z * math.sin(theta)
    return gcj_lon, gcj_lat


def gcj02_to_wgs84(gcj_lon, gcj_lat):
    """GCJ02 → WGS84 坐标转换"""
    a = 6378245.0
    ee = 0.00669342162296594323
    dlat = _transform_lat(gcj_lon - 105.0, gcj_lat - 35.0)
    dlon = _transform_lon(gcj_lon - 105.0, gcj_lat - 35.0)
    radlat = gcj_lat / 180.0 * math.pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * math.pi)
    dlon = (dlon * 180.0) / (a / sqrtmagic * math.cos(radlat) * math.pi)
    wgs_lat = gcj_lat + dlat
    wgs_lon = gcj_lon + dlon
    return wgs_lon, wgs_lat


def _transform_lat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + \
          0.1 * lng * lat + 0.2 * math.sqrt(abs(lng))
    ret += (20.0 * math.sin(6.0 * lng * math.pi) + 20.0 *
            math.sin(2.0 * lng * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * math.pi) + 40.0 *
            math.sin(lat / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * math.pi) + 320 *
            math.sin(lat * math.pi / 30.0)) * 2.0 / 3.0
    return ret


def _transform_lon(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + \
          0.1 * lng * lat + 0.1 * math.sqrt(abs(lng))
    ret += (20.0 * math.sin(6.0 * lng * math.pi) + 20.0 *
            math.sin(2.0 * lng * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * math.pi) + 40.0 *
            math.sin(lng / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * math.pi) + 300.0 *
            math.sin(lng / 30.0 * math.pi)) * 2.0 / 3.0
    return ret


# ============================================================
# xlsx 读取模块（节点坐标）
# ============================================================

def read_xlsx(xlsx_path):
    """
    读取 shanghaidata_x.xlsx，返回节点坐标数据。
    
    返回格式：
    {
      "success": bool,
      "nodes": [{"idx":int, "name":str, "x":float, "y":float, "wgs_lon":float, "wgs_lat":float, ...}],
      "nodeCount": int,
      "columns": [str, ...],
      "xlsx_path": str,
      "xlsx_mtime": str,
      "elapsed_ms": int
    }
    """
    t0 = time.time()
    result = {"success": False, "nodes": [], "nodeCount": 0, "columns": [], "xlsx_path": xlsx_path}

    # 文件存在性检查
    if not os.path.exists(xlsx_path):
        result["error"] = f"文件不存在: {xlsx_path}"
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # 获取文件修改时间
        mtime = os.path.getmtime(xlsx_path)
        from datetime import datetime, timezone, timedelta
        mtime_dt = datetime.fromtimestamp(mtime, tz=timezone(timedelta(hours=8)))
        result["xlsx_mtime"] = mtime_dt.isoformat()

        # 读取表头
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            v = cell.value
            header.append(str(v) if v is not None else "")
        result["columns"] = header

        if not header:
            result["error"] = "xlsx 表头为空"
            return result

        # 建立列名 → 索引映射
        col_map = {}
        for idx, col_name in enumerate(header):
            if col_name:
                col_map[col_name] = idx

        # 读取数据行
        nodes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_list = list(row)
            if not any(v is not None for v in row_list):
                continue  # 跳过空行

            # 提取节点信息
            node = {}
            for col_name, col_idx in col_map.items():
                if col_idx < len(row_list):
                    val = row_list[col_idx]
                    if val is not None:
                        if isinstance(val, (int, float)):
                            node[col_name] = val
                        else:
                            node[col_name] = str(val)
                    else:
                        node[col_name] = None

            # 计算 node_0base（0-based 索引，供数据库匹配使用）
            row_idx = len(nodes)  # 当前已添加的节点数 = 当前行索引（0-based）
            node["node_0base"] = row_idx

            # 提取 BD09 坐标并转换为 WGS84
            x_val = node.get("x")
            y_val = node.get("y")
            if x_val is not None and y_val is not None:
                try:
                    bd_lon = float(x_val)
                    bd_lat = float(y_val)
                    # BD09 → GCJ02 → WGS84
                    gcj_lon, gcj_lat = bd09_to_gcj02(bd_lon, bd_lat)
                    wgs_lon, wgs_lat = gcj02_to_wgs84(gcj_lon, gcj_lat)
                    node["wgs_lon"] = round(wgs_lon, 8)
                    node["wgs_lat"] = round(wgs_lat, 8)
                    # 同时输出 JS 端点期望的字段名
                    node["wgs84_lng"] = round(wgs_lon, 8)
                    node["wgs84_lat"] = round(wgs_lat, 8)
                except (ValueError, TypeError):
                    pass

            nodes.append(node)

        wb.close()

        result["success"] = True
        result["nodes"] = nodes
        result["nodeCount"] = len(nodes)
        result["elapsed_ms"] = int((time.time() - t0) * 1000)

    except ImportError:
        result["error"] = "openpyxl 未安装，请执行: pip install openpyxl"
    except Exception as e:
        result["error"] = f"读取 xlsx 失败: {str(e)}"

    return result


# ============================================================
# CSV 链接矩阵读取模块
# ============================================================

def read_csv_matrix(csv_path):
    """
    读取 link_matrix_shanghai.csv，返回边列表和统计信息。
    
    返回格式：
    {
      "success": bool,
      "edges": [{"source":int, "target":int, "weight":float}],
      "edgeCount": int,
      "matrixShape": [n, m],
      "totalNodes": int,
      "csv_path": str,
      "csv_mtime": str,
      "elapsed_ms": int
    }
    """
    t0 = time.time()
    result = {"success": False, "edges": [], "edgeCount": 0, "matrixShape": [0, 0], "totalNodes": 0, "csv_path": csv_path}

    # 文件存在性检查
    if not os.path.exists(csv_path):
        result["error"] = f"文件不存在: {csv_path}"
        return result

    try:
        from datetime import datetime, timezone, timedelta
        mtime = os.path.getmtime(csv_path)
        mtime_dt = datetime.fromtimestamp(mtime, tz=timezone(timedelta(hours=8)))
        result["csv_mtime"] = mtime_dt.isoformat()

        # 读取 CSV 矩阵
        matrix = []
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)  # 跳过表头行
            for row in reader:
                if len(row) < 2:
                    continue
                values = [float(v) for v in row[1:]]  # 跳过行索引列
                matrix.append(values)

        n = len(matrix)  # 行数（节点数）
        if n == 0:
            result["error"] = "CSV 矩阵为空"
            return result

        m = len(matrix[0]) if n > 0 else 0  # 列数
        result["matrixShape"] = [n, m]
        result["totalNodes"] = n

        # 遍历上三角（i < j），提取 value > 0 的连接边，排除对角线
        edges = []
        for i in range(n):
            for j in range(i + 1, min(m, n)):
                val = matrix[i][j]
                if val > 0:
                    edges.append({
                        "source": i,
                        "target": j,
                        "weight": round(val, 6),
                    })

        result["success"] = True
        result["edges"] = edges
        result["edgeCount"] = len(edges)
        result["elapsed_ms"] = int((time.time() - t0) * 1000)

    except Exception as e:
        result["error"] = f"读取 CSV 失败: {str(e)}"

    return result


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="统一文件读取服务")
    parser.add_argument("--mode", choices=["all", "xlsx", "csv"], default="all",
                        help="读取模式: all=全部, xlsx=仅坐标, csv=仅链接矩阵")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="xlsx 文件路径")
    parser.add_argument("--csv", default=DEFAULT_CSV, help="csv 文件路径")
    args = parser.parse_args()

    output = {"errors": []}

    # 读取 xlsx
    if args.mode in ("all", "xlsx"):
        xlsx_result = read_xlsx(args.xlsx)
        output["xlsx"] = xlsx_result
        if not xlsx_result["success"] and "error" in xlsx_result:
            output["errors"].append(xlsx_result["error"])

    # 读取 csv
    if args.mode in ("all", "csv"):
        csv_result = read_csv_matrix(args.csv)
        output["csv"] = csv_result
        if not csv_result["success"] and "error" in csv_result:
            output["errors"].append(csv_result["error"])

    # 输出 JSON 到 stdout
    sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps(output, ensure_ascii=False, allow_nan=False))


if __name__ == "__main__":
    main()
